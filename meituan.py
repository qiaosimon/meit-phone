# coding=utf-8

import urllib3
import json
import openpyxl
import time
import random
import sys

# 全局计数器
index = 2

# 城市
citys = [{
    'name': '天津',
    'abbr': 'tj',
    'cityid': '40'
}, {
    'name': '西安',
    'abbr': 'xa',
    'cityid': '42'
}, {
    'name': '重庆',
    'abbr': 'cq',
    'cityid': '45'
}, {
    'name': '武汉',
    'abbr': 'wh',
    'cityid': '57'
}, {
    'name': '南京',
    'abbr': 'nj',
    'cityid': '55'
}, {
    'name': '杭州',
    'abbr': 'hz',
    'cityid': '50'
}, {
    'name': '沈阳',
    'abbr': 'sy',
    'cityid': '66'
}, {
    'name': '宁波',
    'abbr': 'nb',
    'cityid': '51'
}, {
    'name': '福州',
    'abbr': 'fz',
    'cityid': '44'
}, {
    'name': '厦门',
    'abbr': 'xm',
    'cityid': '62'
}, {
    'name': '大连',
    'abbr': 'dl',
    'cityid': '65'
}, {
    'name': '徐州',
    'abbr': 'xz',
    'cityid': '119'
}, {
    'name': '苏州',
    'abbr': 'sz',
    'cityid': '80'
}, {
    'name': '成都',
    'abbr': 'cd',
    'cityid': '59'
}, {
    'name': '济南',
    'abbr': 'jn',
    'cityid': '96'
}, {
    'name': '青岛',
    'abbr': 'qd',
    'cityid': '60'
}, {
    'name': '三亚',
    'abbr': 'sanya',
    'cityid': '57'
}, {
    'name': '张家口',
    'abbr': 'zjk',
    'cityid': '125'
}]

now_citys = [{
    'name': '武汉',
    'abbr': 'wh',
    'cityid': '57'
}, {
    'name': '南京',
    'abbr': 'nj',
    'cityid': '55'
}, {
    'name': '杭州',
    'abbr': 'hz',
    'cityid': '50'
}, {
    'name': '南昌',
    'abbr': 'nc',
    'cityid': '83'
}, {
    'name': '郑州',
    'abbr': 'zz',
    'cityid': '73'
}, {
    'name': '长沙',
    'abbr': 'chs',
    'cityid': '70'
}]

now_cates = [{
    'name': '美容美体',
    'cateId': '76'
}, {
    'name': '祛痘',
    'cateId': '20421'
}, {
    'name': '瘦身纤体',
    'cateId': '20422'
}, {
    'name': '医学美容',
    'cateId': '20423'
}]

# 分类
liren_cates = [{
    'name': '养发',
    'cateId': '21394'
}, {
    'name': '美发',
    'cateId': '74'
}, {
    'name': '美容美体',
    'cateId': '76'
}, {
    'name': '美甲美睫',
    'cateId': '75'
}, {
    'name': '瑜伽舞蹈',
    'cateId': '220'
}, {
    'name': '瘦身纤体',
    'cateId': '20422'
}, {
    'name': '瑜伽舞蹈',
    'cateId': '20421'
}]

qingzi_cates = [{
    'name': '亲子活动',
    'cateId': '21284'
}, {
    'name': '亲子会所',
    'cateId': '21283'
}, {
    'name': '儿童乐园',
    'cateId': '20108'
}, {
    'name': '婴儿游泳',
    'cateId': '20514'
}, {
    'name': '新生儿摄影',
    'cateId': '21282'
}, {
    'name': '母婴护理',
    'cateId': '20042'
}]

# 收集到的常用Header
my_headers = [
    "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/537.75.14",
    "Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)",
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) Gecko/20071127 Firefox/2.0.0.11',
    'Opera/9.25 (Windows NT 5.1; U; en)',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)',
    'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)',
    'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.8.0.12) Gecko/20070731 Ubuntu/dapper-security Firefox/1.5.0.12',
    'Lynx/2.8.5rel.1 libwww-FM/2.14 SSL-MM/1.4.1 GNUTLS/1.2.9',
    "Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.7 (KHTML, like Gecko) Ubuntu/11.04 Chromium/16.0.912.77 Chrome/16.0.912.77 Safari/535.7",
    "Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:10.0) Gecko/20100101 Firefox/10.0 "
]


# 判断是否有手机号 粗略判断
def isphone(phone):
    if len(phone) == 11:
        return phone
    else:
        if '/' in phone:
            array = phone.split('/')
            for i in range(2):
                if len(array[i]) == 11:
                    return array[i]


# 请求url中包含城市ID
def geturl(cityid):
    url = "https://apimobile.meituan.com/group/v4/poi/pcsearch/" + cityid
    return url


# 请求头需要城市的简称
def getheaders(abbr):
    headers = {
        "User-Agent": random.choice(my_headers),
        "Referer": "https://" + abbr + ".meituan.com/",
        "Cookie": "uuid=8d3f41b9eee349c9908e.1619505483.1.0.0; _lx_utm=utm_source%3Dbing%26utm_medium%3Dorganic; _lxsdk_cuid=179120bf9b5c8-007d1266fe019e-d7e163f-100200-179120bf9b5c8; mtcdn=K; userTicket=efVfCassnVTuAksNLRXqXAtBTYAEOFsxeqUrXpjV; u=824213835; n=%E4%B8%80%E5%93%A5403; lt=Z62WW2mDrw4qbWcl_oXlJprusjAAAAAAZg0AAGZGBIJYhQJEDcwECzSF2y5W8tB3L__QEhka6mSxg5sk9jc2D01sH3xZuJY6CQKc1w; mt_c_token=Z62WW2mDrw4qbWcl_oXlJprusjAAAAAAZg0AAGZGBIJYhQJEDcwECzSF2y5W8tB3L__QEhka6mSxg5sk9jc2D01sH3xZuJY6CQKc1w; token=Z62WW2mDrw4qbWcl_oXlJprusjAAAAAAZg0AAGZGBIJYhQJEDcwECzSF2y5W8tB3L__QEhka6mSxg5sk9jc2D01sH3xZuJY6CQKc1w; lsu=; token2=Z62WW2mDrw4qbWcl_oXlJprusjAAAAAAZg0AAGZGBIJYhQJEDcwECzSF2y5W8tB3L__QEhka6mSxg5sk9jc2D01sH3xZuJY6CQKc1w; unc=%E4%B8%80%E5%93%A5403; ci=83; rvct=83%2C73%2C70%2C50; firstTime=1619507605826; _lxsdk_s=179120bf9b5-6a8-76b-b93%7C%7C28"
    }
    return headers


# 请求参数 offset 偏移量  limit 大小  cateid 分类id
def getdate(offset, limit, cateId):
    data = {
        "uuid": "90478bb8bf3b4047b234.1618798623.1.0.0",
        "userid": "1",  # 用户ID需要修改
        "limit": limit,
        "offset": offset,
        "cateId": cateId,
        "token": "im2rm2IpGGak0SngipFRGbIzLooAAAAARg0AAI8fCS-046-PkfJbyoZlu4wHogUiw9X2lne6jAusBowI-vtU8RDQIzP83MlSr37Isw",
        "areaId": -1
    }
    return data


# 保存数据到excel
def saveExcel(cityname, catename, response):
    result = json.loads(response.data.decode('utf-8'))
    searchResult = result.get('data').get('searchResult')
    # print(searchResult)
    for temp in searchResult:
        title = temp.get('title')
        address = temp.get('address')
        phone = temp.get('phone')
        avgscore = temp.get('avgscore')  # 评分
        comments = temp.get('comments')  # 评论数
        historyCouponCount = temp.get('historyCouponCount')  # 历史成交单数
        # 判断是否有手机号
        phonestr = isphone(phone)
        if not phonestr is None:
            global index

            print(str(
                index) + '***' + cityname + '***' + catename + '***' + title + '***' + address + '***' + phonestr + '***' + str(
                avgscore) + '***' + str(comments) + '***' + str(historyCouponCount))
            # 调整行高
            sheet.row_dimensions[index].height = 25
            # 保存数据
            sheet.cell(index, 1, cityname)
            sheet.cell(index, 2, catename)
            sheet.cell(index, 3, title)
            sheet.cell(index, 4, address)
            sheet.cell(index, 5, phonestr)
            sheet.cell(index, 6, avgscore)
            sheet.cell(index, 7, comments)
            sheet.cell(index, 8, historyCouponCount)
            index = index + 1

    # 保存数据
    workbook.save('D:/美容院.xlsx')


# 表格
workbook = openpyxl.Workbook()
sheet = workbook.active
# 调整列宽
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 5
sheet.column_dimensions['C'].width = 50
sheet.column_dimensions['D'].width = 60
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 5
sheet.column_dimensions['G'].width = 6
sheet.column_dimensions['H'].width = 11

# 标题
sheet.cell(1, 1, '城市')
sheet.cell(1, 2, '分类')
sheet.cell(1, 3, '店铺名')
sheet.cell(1, 4, '地址')
sheet.cell(1, 5, '电话')
sheet.cell(1, 6, '评分')
sheet.cell(1, 7, '评论数')
sheet.cell(1, 8, '历史成交单数')

for city in now_citys:
    cityname = city.get('name')
    abbr = city.get('abbr')
    cityid = city.get('cityid')

    # 创建http请求对象
    http = urllib3.PoolManager()
    for cate in now_cates:
        catename = cate.get('name')
        cateid = cate.get('cateId')

        limit = 100
        offset = 0
        url = geturl(cityid)
        fields = getdate(offset, limit, cateid)
        headers = getheaders(abbr)
        # 发送请求
        response = http.request('GET', url=geturl(cityid), fields=fields, headers=headers)
        print(response.status)
        if response.status == 200:
            print('数据正常')
            saveExcel(cityname, catename, response)
        else:
            print('******************error! system exit************************')
            sys.exit()
        sleeptime = random.randint(1, 15)
        print('睡眠' + str(sleeptime) + '秒')  # 防止请求过多被封账号
        time.sleep(sleeptime)
